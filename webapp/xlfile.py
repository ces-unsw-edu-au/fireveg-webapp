import click
from flask import current_app, g
from flask.cli import with_appcontext

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, PatternFill, Border, Font # Side, Alignment, Protection,
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd

import datetime

# common stuff
cent_align=Alignment(horizontal='center', vertical='center', wrap_text=False)
wrap_align=Alignment(horizontal='left', vertical='top', wrap_text=True)
sheet_colors = {"instructions": "1072BA" , "intro": "1072BA" , "entry": "10BA72", "summary": "5AFF5A", "addentry": "20CA82", "default":"505050"}
fontSmall = Font(size = "9")
table_style={
    "Instructions":TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True, showLastColumn=False, showRowStripes=True, showColumnStripes=False),
     "Contributor": TableStyleInfo(name="TableStyleMedium18", showFirstColumn=True, showLastColumn=False, showRowStripes=False, showColumnStripes=False),
     "Lists": TableStyleInfo(name="TableStyleMedium14", showFirstColumn=True, showLastColumn=False, showRowStripes=False, showColumnStripes=False),
     "Info":  TableStyleInfo(name="TableStyleMedium14", showFirstColumn=True, showLastColumn=False, showRowStripes=False, showColumnStripes=False),
     "Vocabularies": TableStyleInfo(name="TableStyleMedium14", showFirstColumn=True, showLastColumn=False, showRowStripes=False, showColumnStripes=False),
     "Entry": TableStyleInfo(name="TableStyleMedium18", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
     }

#####
## Content:
## Content for these functions is declared in xlcontent.py and saved to pickle files in folder content
#####


#####
## This is the function to create a detailed output workbook, each row is a record
#####
def create_list_records_xl(
traitsummary=None, referencelist=None, traitlist=None, wsheets=None, info=None, supporters=None, description=None):
    wb = Workbook()
    ws = wb.active

    for item in wsheets:
        if "active" in item.keys():
            ws = wb.active
            ws.title = item['title']
        else:
            ws = wb.create_sheet(item['title'])
        for k in item['colWidths']:
            for j in k[0]:
                ws.column_dimensions[j].width = k[1]
                ws.sheet_properties.tabColor = sheet_colors[item["tabColor"]]

    ## About
    ws = wb["About"]
    k = 1
    for row in info:
        ws.cell(k,1,value=row)
        ws.cell(k,1).alignment=wrap_align
        k=k+1
    ## check these if info was updated/changed
    ws.cell(1,1).style='Title'
    ws.cell(5,1).hyperlink='https://www.unsw.edu.au/research/ecosystem'
    ws.cell(5,1).style='Hyperlink'
    # Disclaimer, check these if info was updated/changed
    ws.cell(8,1).font=Font(color="FF0000", bold=True,italic=False)
    ws.cell(9,1).font=Font(color="FF0000", italic=True)
    k=k+2
    ws.cell(k-1,1,value="This work has been supported by:")
    for item in supporters:
        cell=ws.cell(k,1)
        cell.value=item['institution']
        cell.hyperlink=item['url']
        cell.style = "Hyperlink"
        k=k+1
    k=k+2
    for row in description['about']:
        ws.cell(k,1,value=row)
        ws.cell(k,1).alignment=wrap_align
        k=k+1
    ws.protection.sheet = True

    ## Trait Description
    ws = wb["Trait description"]
    k=1
    for row in description["traits"]:
        ws.cell(k,3,value=row)
        ws.cell(k,3).alignment=wrap_align
        k=k+1

    tab_begin=k
    ws.append(["Trait Code", "Trait Name", "Description", "Type", "Life stage", "Life history process", "Data migration"])
    k=k+1

    for row in traitlist:
        j=1
        for key in ["code","name","description","value_type","life_stage","life_history_process","priority"]:
            val=row[key]
            ws.cell(row=k, column=j, value=val)
            j=j+1
        k=k+1

    for j in range(tab_begin,ws.max_row+1):
        ws.cell(j,3).alignment=wrap_align

    tab = Table(displayName="TraitInformation", ref="A{}:G{}".format(tab_begin,ws.max_row))

    tab.tableStyleInfo = table_style["Info"]
    ws.add_table(tab)
    ws.protection.sheet = True

    ## Summary

    ws = wb["Summary"]

    colnames = ['scientific name','current code (BioNET)',
                'original name (as entered)','CAPS code (old)',
                'trait code','trait name','norm value',
                'best','lower','upper',
                'method of estimation',
                'weight','source ref','other ref','DB link']
    ws.append(colnames)

    rows = traitsummary.sort_values(by =['scientific name','trait code']).to_dict(orient="records")


    for r_idx, row in enumerate(rows, 2):

        ws.cell(row=r_idx, column=1, value=row['scientific name'])
        ws.cell(r_idx,1).font  = Font(italic=True)

        ws.cell(row=r_idx, column=2, value=row['current code (BioNET)'])
        if row['original name'] != row['scientific name']:
            ws.cell(row=r_idx, column=3, value=row['original name'])
            ws.cell(r_idx,3).font  = Font(italic=True, color="110000")
            ws.cell(row=r_idx, column=4, value=row['CAPS code'])
        ws.cell(row=r_idx, column=5, value=row['trait code'])
        ws.cell(row=r_idx, column=6, value=row['trait name'])

        ws.cell(row=r_idx, column=11, value=row['method'])
        ws.cell(row=r_idx, column=12, value=row['weight'])
        ws.cell(row=r_idx, column=13, value=row['source ref'])
        if row['other ref'] is not None:
            oref="; ".join(row['other ref'])
            ws.cell(row=r_idx, column=14, value=oref)

        if isinstance(row['norm value'],str):
            val=row['norm value']
        elif row['norm value'] is None:
            val="(data input ERROR)"
        else:
            triplet=row['norm value']
            k=7
            for j in triplet:
                k=k+1
                if j is not None:
                    ws.cell(row=r_idx, column=k, value=j)
            if triplet[0] is not None:
                if triplet[1] is None and triplet[2] is None:
                    val=triplet[0]
                else:
                    val = "%s (%s -- %s)" % tuple(triplet)
            else:
                if triplet[1] is None:
                    if triplet[2] is None:
                        val="(data input ERROR)"
                    else:
                        val = "<%s" % triplet[2]
                elif triplet[2] is None:
                    val = ">%s" % triplet[1]
                else:
                    val = "(%s -- %s)" % (triplet[1],triplet[2])
        ws.cell(row=r_idx, column=7, value=val)

        val = "trait:%s / sp code:%s / record id:%s" % (row['trait code'],row['CAPS code'],row['recordid'])
        url = "http://13.54.3.205/traits/%s/%s" % (row['trait code'],row['CAPS code'])
        cell=ws.cell(row=r_idx, column=15, value=val)
        cell.hyperlink=url
        cell.style='Hyperlink'

        for j in (2,4,5,7,12):
            ws.cell(r_idx,j).alignment=cent_align
        for j in (11,13,14,15):
            ws.cell(r_idx,j).font = fontSmall
            ws.cell(r_idx,j).alignment=wrap_align


    tab = Table(displayName="Summary", ref="A1:{}{}".format(get_column_letter(15),r_idx))
    tab.tableStyleInfo = table_style["Lists"]
    ws.add_table(tab)


    ## References
    ws = wb["References"]

    k=1

    for row in description["references"]:
        ws.cell(k,2,value=row)
        ws.cell(k,2).alignment=wrap_align
        k=k+1


    ws.append(["Reference code", "Reference information"])

    for row in referencelist:
        ws.append(row)

    #ws.max_row
    for j in range(k+1,ws.max_row+1):
        ws.cell(j,2).alignment=wrap_align
        ws.cell(j,2).font = fontSmall

    tab = Table(displayName="ReferenceInformation", ref="A{}:B{}".format(k,ws.max_row))

    tab.tableStyleInfo = table_style["Lists"]
    ws.add_table(tab)
    ws.protection.sheet = True

    ## Finalise and return
    return wb

#####
## This is the function to create a summary output report
#####
def create_output_xl(traitsummary=None, referencelist=None, traitlist=None, info=None, wsheets=None, description=None, supporters=None):
    wb = Workbook()
    ws = wb.active

    for item in wsheets:
        if "active" in item.keys():
            ws = wb.active
            ws.title = item['title']
        else:
            ws = wb.create_sheet(item['title'])
        for k in item['colWidths']:
            for j in k[0]:
                ws.column_dimensions[j].width = k[1]
        ws.sheet_properties.tabColor = sheet_colors[item["tabColor"]]

    ## About
    ws = wb["About"]
    k = 1
    for row in info:
        ws.cell(k,1,value=row)
        ws.cell(k,1).alignment=wrap_align
        k=k+1
    ## check these if info was updated/changed
    ws.cell(1,1).style='Title'
    ws.cell(5,1).hyperlink='https://www.unsw.edu.au/research/ecosystem'
    ws.cell(5,1).style='Hyperlink'
    # Disclaimer, check these if info was updated/changed
    ws.cell(8,1).font=Font(color="FF0000", bold=True,italic=False)
    ws.cell(9,1).font=Font(color="FF0000", italic=True)
    k=k+2
    ws.cell(k-1,1,value="This work has been supported by:")
    for item in supporters:
        cell=ws.cell(k,1)
        cell.value=item['institution']
        cell.hyperlink=item['url']
        cell.style = "Hyperlink"
        k=k+1
    k=k+2
    for row in description['about']:
        ws.cell(k,1,value=row)
        ws.cell(k,1).alignment=wrap_align
        k=k+1
    ws.protection.sheet = True

    ## Trait Description
    ws = wb["Trait description"]
    k=1
    for row in description['traits']:
        ws.cell(k,3,value=row)
        ws.cell(k,3).alignment=wrap_align
        k=k+1
    ws.append(["Trait Code", "Trait Name", "Description", "Type", "Life stage", "Life history process", "Data migration"])
    for row in traitlist:
        ws.append(row)
    #ws.max_row
    for j in range(k,ws.max_row+1):
        ws.cell(j,3).alignment=wrap_align
    tab = Table(displayName="TraitInformation", ref="A{}:G{}".format(k,ws.max_row))
    tab.tableStyleInfo = table_style["Info"]
    ws.add_table(tab)
    ws.protection.sheet = True

    ## Summary
    ws = wb["Summary"]
    ws.append(['Species','Code','surv1','surv4','germ1','germ8','rect2','repr2','repr3','repr3a','disp1','Original Species name(s) used','Import/Entry sources','Indirect sources'])
    rows = dataframe_to_rows(traitsummary[['Species','Code','surv1','surv4','germ1','germ8','rect2','repr2','repr3','repr3a','disp1','orig_species','main_refs','orig_refs']],index=False, header=False)
    #rows = dataframe_to_rows(traitsummary,index=False, header=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        for k in (12,13,14):
            ws.cell(r_idx,k).alignment=wrap_align
            ws.cell(r_idx,k).font = fontSmall
    tab = Table(displayName="Summary", ref="A1:{}{}".format(get_column_letter(c_idx),r_idx))
    tab.tableStyleInfo = table_style["Lists"]
    ws.add_table(tab)

    ## References
    ws = wb["References"]
    k=1

    for row in description['references']:
        ws.cell(k,2,value=row)
        ws.cell(k,2).alignment=wrap_align
        k=k+1
    ws.append(["Reference code", "Reference information"])
    for row in referencelist:
        ws.append(row)
    #ws.max_row
    for j in range(k+1,ws.max_row+1):
        ws.cell(j,2).alignment=wrap_align
        ws.cell(j,2).font = fontSmall
    tab = Table(displayName="ReferenceInformation", ref="A{}:B{}".format(k,ws.max_row))
    tab.tableStyleInfo = table_style["Lists"]
    ws.add_table(tab)
    ws.protection.sheet = True

    ## Finalise and return
    return wb

#####
## This is the function to create a data entry document to download
#####
def create_input_xl(contactinfo=None, specieslist=None, referencelist=None, traitlist=None, vocabularies=None, methods_vocabularies=None, wsheets=None, instructions=None, links=None):
    wb = Workbook()
    ws = wb.active

    for item in wsheets:
        if "active" in item.keys():
            ws = wb.active
            ws.title = item['title']
        else:
            ws = wb.create_sheet(item['title'])
        for k in item['colWidths']:
            for j in k[0]:
                ws.column_dimensions[j].width = k[1]
        ws.sheet_properties.tabColor = sheet_colors[item["tabColor"]]

    ## instructions
    ws = wb["Instructions"]


    ws.append(["Step", "Instructions","Links"])
    for k in range(len(instructions)):
        ws.cell(k+2,1).value=k+1
        ws.cell(k+2,1).alignment=cent_align
        ws.cell(k+2,2).value=instructions[k]
        ws.cell(k+2,2).alignment=wrap_align
        if links[k] is not None:
            cell=ws.cell(k+2,3)
            cell.value=links[k][1]
            cell.hyperlink=links[k][0]
            cell.style = "Hyperlink"
    tab = Table(displayName="Instructions", ref="A1:C{}".format(len(instructions)+1))
    tab.tableStyleInfo = table_style["Instructions"]
    ws.add_table(tab)

    ## contributor
    ws = wb['Contributor']
    if contactinfo is None:
        contactinfo = {
            'Name': """ Your name """,
            'Affiliation': """ Your institution """,
            'Contact': """ e-mail or phone """
        }
    ws.append(["Field", "Your response"])
    for key in contactinfo.keys():
        ws.append((key,contactinfo[key]))
    tab = Table(displayName="Contributor", ref="A1:B4")
    tab.tableStyleInfo = table_style["Contributor"]
    ws.add_table(tab)

    ## Species list
    if specieslist is not None:
        ws = wb["Species list"]
        ws.append(["Scientific Name","Code","Family", "Genus", "Scientific Name ID in BioNET", "current Scientific Name Code", "Current Scientific Name", "Current Vernacular Name", "is Current?"])
        for row in specieslist:
            ws.append(row)
        tab = Table(displayName="SpeciesList", ref="A1:B{}".format(ws.max_row))
        tab.tableStyleInfo = table_style["Lists"]
        ws.add_table(tab)
        ws.protection.sheet = True

    ## Reference list
    if referencelist is not None:
        ws = wb["References"]
        ws.append(["Code", "Full reference"])
        for row in referencelist:
            ws.append(row)
        for k in range(2,ws.max_row+1):
            ws.cell(k,2).alignment=wrap_align
        tab = Table(displayName="References", ref="A1:B{}".format(ws.max_row))
        tab.tableStyleInfo = table_style["Lists"]
        ws.add_table(tab)

    ## Trait description
    if traitlist is not None:
        ws = wb["Trait description"]
        ws.append(["Trait Code", "Trait Name", "Description", "Type", "Life stage", "Life history process", "Priority"])
        for row in traitlist:
            ws.append(row)
        for k in range(2,ws.max_row+1):
            ws.cell(k,3).alignment=wrap_align
        tab = Table(displayName="TraitInformation", ref="A1:G{}".format(ws.max_row))
        tab.tableStyleInfo = table_style["Info"]
        ws.add_table(tab)
        ws.protection.sheet = True

    ## Vocabularies
    if vocabularies is not None:
        ws = wb["Vocabularies"]
        k=1
        for record in vocabularies:
            ws.cell(row=k,column=1,value="Lookup table for trait %s" % record['code'])
            k=k+1
            tab_first_row=k
            ws.cell(row=k,column=1,value="Valid values")
            ws.cell(row=k,column=2,value="Description")
            vocab=record['vocab']
            for key in vocab.keys():
                k=k+1
                ws.cell(row=k,column=1,value=key)
                ws.cell(row=k,column=2,value=vocab[key])
                ws.cell(row=k,column=2).alignment=wrap_align
            tab_last_row=k
            tab = Table(displayName="lookup_%s" % record['code'], ref="A{}:B{}".format(tab_first_row,tab_last_row))
            tab.tableStyleInfo = table_style["Vocabularies"]
            ws.add_table(tab)
            k=k+2
        ws.protection.sheet = True

    ## methods Vocabularies
    if methods_vocabularies is not None:
        ws = wb["Vocabularies for methods"]
        k=1
        for record in methods_vocabularies:
            ws.cell(row=k,column=1,value="Available methods for trait %s" % record['code'])
            k=k+1
            tab_first_row=k
            ws.cell(row=k,column=1,value="Valid values")
            ws.cell(row=k,column=2,value="Description")
            vocab=record['vocab']
            for key in vocab.keys():
                k=k+1
                ws.cell(row=k,column=1,value=key)
                ws.cell(row=k,column=2,value=vocab[key])
                ws.cell(row=k,column=2).alignment=wrap_align
            tab_last_row=k
            tab = Table(displayName="method_%s" % record['code'], ref="A{}:B{}".format(tab_first_row,tab_last_row))
            tab.tableStyleInfo = table_style["Vocabularies"]
            ws.add_table(tab)
            k=k+2
        ws.protection.sheet = True

    ## Data Entry
    ws = wb["Data entry"]
    nrows = 200
    hdr=["Main source", "Original sources", "Original species name", "Species code", "Species name", "Trait code", "Trait name","Trait type","Raw value", "Norm value", "Best", "Lower", "Upper", "Method of estimation","Notes"]
    ws.append(hdr)
    dv_ref = DataValidation(type="list",
                        formula1="""=INDIRECT("References[Code]")""",
                        allow_blank=True)

    dv_fuzzy = DataValidation(type="decimal",
                        operator="greaterThanOrEqual",
                        formula1=0)

    dv_trait = DataValidation(type="list",
                        formula1="""=INDIRECT("TraitInformation[Trait Code]")""")

    dv_vvalue = DataValidation(type="list",
                        formula1="""=INDIRECT(CONCATENATE("lookup_",$F2,"[Valid values]"))""")
    dv_mvalue = DataValidation(type="list",
                    formula1="""=INDIRECT(CONCATENATE("method_",$F2,"[Valid values]"))""")

    # custom error message
    dv_ref.error ='Your entry is not in the list'
    dv_ref.errorTitle = 'Invalid Entry'
    dv_trait.error ='Your entry is not in the list'
    dv_trait.errorTitle = 'Invalid Entry'

    # custom prompt message
    dv_ref.prompt = 'Please select from the list of references'
    dv_ref.promptTitle = 'List Selection'
    dv_vvalue.prompt = """For categorical traits, please select trait first and then select one value from the dropdown list, otherwise leave blank.
    For quantitative traits, leave blank and fill Best/Lower/Upper columns."""
    dv_vvalue.promptTitle = 'Accepted values for trait'
    dv_mvalue.prompt = """Please select a valid method for this trait from the dropdown list. Leave blank if no options available."""
    dv_mvalue.promptTitle = 'Accepted values for method'

    # add validation ranges
    dv_ref.add("A2:A{}".format(nrows+1))
    dv_trait.add("F2:F{}".format(nrows+1))
    dv_fuzzy.add("K2:M{}".format(nrows+1))
    dv_vvalue.add("J2:J{}".format(nrows+1))
    dv_mvalue.add("N2:N{}".format(nrows+1))

    ## add to sheet
    ws.add_data_validation(dv_vvalue)
    ws.add_data_validation(dv_mvalue)
    ws.add_data_validation(dv_ref)
    ws.add_data_validation(dv_trait)
    ws.add_data_validation(dv_fuzzy)

    for row in range(2,nrows+2):
        cell=ws.cell(row=row,column=7)
        cell.value="""=VLOOKUP($F{}, INDIRECT("TraitInformation"), 2, FALSE)""".format(cell.row)
        cell=ws.cell(row=row,column=8)
        cell.value="""=VLOOKUP($F{}, INDIRECT("TraitInformation"), 4, FALSE)""".format(cell.row)
        cell=ws.cell(row=row,column=4)
        cell.value="""=VLOOKUP($C{}, INDIRECT("SpeciesList"), 2, FALSE)""".format(cell.row)
        cell=ws.cell(row=row,column=5)
        cell.value="""=VLOOKUP($C{}, INDIRECT("SpeciesList"), 1, FALSE)""".format(cell.row)

    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(fill=red_fill)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = ['$H2="categorical"']
    ws.conditional_formatting.add("K2:M{}".format(nrows+1), r)

    r2 = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r2.formula = ['$H2="numerical"']
    ws.conditional_formatting.add("J2:J{}".format(nrows+1), r2)

    #r3 = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    # which formula allows to test if table exists?
    #r3.formula = ['=ISNUMBER(ROWS(INDIRECT(CONCATENATE("method_",$F2))))'] # formula for counting rows in a table
    #ws.conditional_formatting.add("N2:N{}".format(nrows+1), r2)

    cell=ws.cell(row=nrows+1,column=len(hdr))
    tab = Table(displayName="DataEntry", ref="A1:{}{}".format(cell.column_letter,nrows+1))
    tab.tableStyleInfo = table_style["Entry"]
    ws.add_table(tab)

    return wb
