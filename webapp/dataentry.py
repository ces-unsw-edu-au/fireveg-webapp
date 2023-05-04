from flask import (
    Blueprint, flash, g, redirect, render_template, current_app, request, url_for, send_file
)
import pandas as pd
import io

from werkzeug.utils import secure_filename
from werkzeug.exceptions import abort

from webapp.auth import login_required
from webapp.db import get_db
from webapp.xlfile import create_input_xl
from webapp.pg import get_pg_connection
from psycopg2.extras import DictCursor
import openpyxl
import os
import datetime

bp = Blueprint('dataentry', __name__, url_prefix='/data-entry')


def make_tree(path):
    tree = dict(name=os.path.basename(path), children=[])
    try: lst = os.listdir(path)
    except OSError:
        pass #ignore errors
    else:
        for name in lst:
            fn = os.path.join(path, name)
            if os.path.isdir(fn):
                tree['children'].append(make_tree(fn))
            else:
                tree['children'].append(dict(name=name))
    return tree

@bp.route('/', methods=('GET', 'POST'))
@login_required
def howto():
    #if request.method == 'POST':
    #    destination=request.form['destination']
        # # Quick option, for small instances
    #    if destination=='data entry':
    #         return send_file(current_app.config['DATAENTRY'],
    #         attachment_filename="fire-ecology-traits-data-entry-form.xlsx",
    #         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #         cache_timeout=0)

    #    elif destination=='field proforma':
    #         return send_file(current_app.config['PROFORMA'],         attachment_filename="fire-ecology-fieldwork-proforma.docx", mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document', cache_timeout=0)
    #    else:
    #        return "file not found!"
    #else:
    s3list=[
    'https://fireveg-db.s3.ap-southeast-2.amazonaws.com/output-report/fireveg-trait-records-model.xlsx',
    'https://fireveg-db.s3.ap-southeast-2.amazonaws.com/output-report/fireveg-field-report-model.xlsx'
        ]
    return render_template('data-entry.html', the_title="Data Entry",s3file=s3list)    


@bp.route('/upload/<destination>', methods=('GET', 'POST'))
@bp.route('/upload', defaults={'destination': None}, methods=('GET', 'POST'))
@login_required
def upload_file(destination):
    if request.method == 'POST':
        #print(request.files['file'])
        print(request)
        f = request.files['file']
        filename= datetime.datetime.now().strftime('%Y%m%d_%H%M%S_') + secure_filename(f.filename)
        upload_path = os.path.join(current_app.config['UPLOAD_FOLDER'],
        request.form['destination'],
        filename)
        f.save(upload_path)
        path = current_app.config['UPLOAD_FOLDER']
        return render_template('data-entry/show-dir.html', tree=make_tree(path), uploadedfile=filename)
        #data_xls = pd.read_excel(f)
        #return data_xls.to_html()
    return render_template('data-entry/upload.html',destination=destination)

## create and download excel file for data entry:

@bp.route('/download/<destination>', methods=('GET', 'POST'))
@bp.route('/download', defaults={'destination': None}, methods=('GET', 'POST'))
@login_required
def download_file(destination):
    if request.method == 'POST':
        destination=request.form['destination']
        # Quick option, for small instances
        if destination=='data entry':
            #return send_file(current_app.config['DATAENTRY'],         attachment_filename="fire-ecology-traits-data-entry-form.xlsx",             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', cache_timeout=0)
            print(destination)
        elif destination=='field proforma':
            print("hello")
            #return send_file(current_app.config['PROFORMA'],         attachment_filename="fire-ecology-fieldwork-proforma.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', cache_timeout=0)
        else:
            print("bye")
        # If we can afford a better instance, we can use this time/resource consumming alternative:

        #contactinfo = request.form
        #wb = openpyxl.load_workbook(current_app.config['DATAENTRY'])
        #if contactinfo is not None:
        #     ws = wb['Contributor']
        #     ws.cell(row=2,column=2,value=contactinfo['Name'])
        #     ws.cell(row=3,column=2,value=contactinfo['Affiliation'])
        #     ws.cell(row=4,column=2,value=contactinfo['Contact'])
        #
        # excel_stream = io.BytesIO()
        # wb.save(excel_stream)
        # excel_stream.seek(0)  # go to the beginning of the stream
        # return send_file(
        #         excel_stream,
        #         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        #         attachment_filename="fire-ecology-traits-data-entry-form.xlsx",
        #         as_attachment=True,
        #         cache_timeout=0)
    else:
        return render_template('data-entry/download.html')
